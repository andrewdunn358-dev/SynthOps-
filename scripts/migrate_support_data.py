#!/usr/bin/env python3
"""
migrate_support_data.py
-----------------------
Imports historical data from Monthly_Support_Count.xlsx and Monthly_Client_Changes.xlsx
into SynthOps via the /api/support/import endpoint.

Usage:
    pip install openpyxl requests
    python3 scripts/migrate_support_data.py \
        --support-count Monthly_Support_Count.xlsx \
        --changes Monthly_Client_Changes.xlsx \
        --url http://localhost:8001 \
        --token <your-jwt-token>

    # Dry run (prints JSON, does not POST):
    python3 scripts/migrate_support_data.py ... --dry-run
"""

import argparse
import json
import uuid
import sys
from datetime import datetime, timezone
from collections import defaultdict

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl requests")
    sys.exit(1)


# ── Column layout for the Support Count sheet ──────────────────────────────
# The spreadsheet has a 3-row header (rows 1-3). Row 3 is 'Company' row.
# Columns vary slightly between sheets so we map by known names.

SUPPORT_TYPE_COL = "support type"
BITDEFENDER_COL = "bitdefender"
BACKUP_COL = "backup"

# Device count columns
DEVICE_COLS = {
    "physical server": "Physical Server",
    "virtual server": "Virtual Server",
    "laptop": "Laptop",
    "tablets": "Tablet",
    "desktop": "Desktop",
}

# Onsite device columns
ONSITE_COLS = {
    "router/fw": "Router/FW",
    "switch": "Switch",
    "printer": "Printer",
    "nas": "NAS",
    "voip": "VoIP",
    "wifi ap": "WiFi AP",
}

# Office 365 columns
O365_COLS = {
    "standard": "O365 Standard",
    "standard (monthly commitment)": "O365 Standard Monthly",
    "apps": "O365 Apps",
    "basic": "O365 Basic",
    "exchange online plan 1": "Exchange Online Plan 1",
    "exchange online plan 2": "Exchange Online Plan 2",
    "enterprise": "O365 Enterprise",
    "net sheriff": "Net Sheriff",
    "message labs": "Message Labs",
    "adobe acrobat": "Adobe Acrobat",
}

OTHER_COLS = {
    "broadband": "Broadband",
    "hosting/email": "Hosting/Email",
    "domain name": "Domain Names",
}

# Build reverse lookup: canonical product name -> category
PRODUCT_CATEGORIES = {}
for v in DEVICE_COLS.values():
    PRODUCT_CATEGORIES[v] = "devices"
for v in ONSITE_COLS.values():
    PRODUCT_CATEGORIES[v] = "onsite"
for v in O365_COLS.values():
    PRODUCT_CATEGORIES[v] = "office365"
PRODUCT_CATEGORIES["Bitdefender"] = "security"
PRODUCT_CATEGORIES["Backup Storage"] = "backup"
PRODUCT_CATEGORIES["Broadband"] = "connectivity"
PRODUCT_CATEGORIES["Hosting/Email"] = "hosting"
PRODUCT_CATEGORIES["Domain Names"] = "hosting"
PRODUCT_CATEGORIES["Net Sheriff"] = "other"
PRODUCT_CATEGORIES["Message Labs"] = "other"
PRODUCT_CATEGORIES["Adobe Acrobat"] = "other"

PRODUCT_UNITS = {
    "Backup Storage": "gb",
    "Broadband": "text",
    "Hosting/Email": "yes/no",
    "Domain Names": "text",
}


def parse_month_label(sheet_name: str) -> str:
    """Convert sheet name like 'Aug 23', 'January 2026' to YYYY-MM"""
    name = sheet_name.strip()
    formats = [
        "%b %y", "%b %Y", "%B %y", "%B %Y",
        "%b%y", "%b%Y", "%B%y", "%B%Y",
        "%b. %Y", "%b.%Y",
    ]
    # normalise spacing
    name_clean = " ".join(name.split())
    for fmt in formats:
        try:
            dt = datetime.strptime(name_clean, fmt)
            return dt.strftime("%Y-%m")
        except ValueError:
            continue
    # Try abbreviated like 'JAN25' -> 'JAN 25'
    import re
    m = re.match(r"([A-Za-z]+)[\s\-]?(\d{2,4})$", name_clean)
    if m:
        mon, yr = m.group(1), m.group(2)
        if len(yr) == 2:
            yr = "20" + yr
        for fmt in ["%b %Y", "%B %Y"]:
            try:
                dt = datetime.strptime(f"{mon} {yr}", fmt)
                return dt.strftime("%Y-%m")
            except ValueError:
                continue
    print(f"  WARNING: Could not parse sheet month '{sheet_name}', skipping")
    return None


def get_col_map(header_rows):
    """
    Build a column-index map from the merged 3-row header.
    Returns dict: canonical_lower -> col_index
    """
    col_map = {}
    # Row 1 (index 0): section headers like 'Device Count', 'Onsite Devices', 'OFFICE 365'
    # Row 2 (index 1): sub-headers like 'Physical Server', 'Virtual Server', etc.
    # Row 3 (index 2): 'Company' row — this is our data start

    if len(header_rows) < 3:
        return col_map

    row2 = header_rows[1]   # sub-headers
    row3 = header_rows[2]   # 'Company' row

    # Use row2 + row3 combined: prefer row2 if set, otherwise row3
    for i, val in enumerate(row2):
        if val and str(val).strip():
            col_map[str(val).strip().lower()] = i
    for i, val in enumerate(row3):
        if val and str(val).strip():
            key = str(val).strip().lower()
            if key not in col_map:
                col_map[key] = i
    return col_map


def clean_value(val):
    """Return None for empty/whitespace cells"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None", "-", "N/A", "NA", "n/a"):
        return None
    return s


def parse_backup_gb(val):
    """Extract numeric GB value from strings like '142GB', '450GB'"""
    if val is None:
        return None
    import re
    m = re.match(r"([\d.]+)\s*[Gg][Bb]?", str(val))
    if m:
        return float(m.group(1))
    return clean_value(val)


def normalise_company(name: str) -> str:
    return " ".join(str(name).split()).strip().lower()


def parse_support_count(wb) -> tuple[dict, dict]:
    """
    Returns:
        products_seen: set of product names encountered
        monthly_data: { YYYY-MM: { company_name_lower: { ...fields } } }
    """
    products_seen = set()
    monthly_data = {}

    for sheet_name in wb.sheetnames:
        month_key = parse_month_label(sheet_name)
        if not month_key:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # First 3 rows are headers
        col_map = get_col_map(rows[:3])
        data_rows = rows[3:]

        month_clients = {}

        for row in data_rows:
            company = clean_value(row[0] if row else None)
            if not company:
                continue
            # Skip the example row
            if "example" in company.lower():
                continue

            entry = {"support_type": None, "products": {}, "remarks": None}

            # Support type
            idx = col_map.get(SUPPORT_TYPE_COL)
            if idx is not None and idx < len(row):
                entry["support_type"] = clean_value(row[idx])

            # Bitdefender
            idx = col_map.get(BITDEFENDER_COL)
            if idx is not None and idx < len(row):
                val = clean_value(row[idx])
                if val and val.lstrip("-").isdigit():
                    products_seen.add("Bitdefender")
                    entry["products"]["Bitdefender"] = int(float(val))

            # Backup storage
            idx = col_map.get(BACKUP_COL)
            if idx is not None and idx < len(row):
                bval = parse_backup_gb(row[idx])
                if bval is not None:
                    products_seen.add("Backup Storage")
                    entry["products"]["Backup Storage"] = bval

            # Device columns
            for col_key, product_name in DEVICE_COLS.items():
                idx = col_map.get(col_key)
                if idx is not None and idx < len(row):
                    val = clean_value(row[idx])
                    if val and str(val).lstrip("-").replace(".", "").isdigit():
                        products_seen.add(product_name)
                        entry["products"][product_name] = int(float(val))

            # Onsite columns
            for col_key, product_name in ONSITE_COLS.items():
                idx = col_map.get(col_key)
                if idx is not None and idx < len(row):
                    val = clean_value(row[idx])
                    if val and str(val).lstrip("-").replace(".", "").isdigit():
                        products_seen.add(product_name)
                        entry["products"][product_name] = int(float(val))

            # Office 365 columns
            for col_key, product_name in O365_COLS.items():
                idx = col_map.get(col_key)
                if idx is not None and idx < len(row):
                    val = clean_value(row[idx])
                    if val and str(val).lstrip("-").replace(".", "").isdigit():
                        products_seen.add(product_name)
                        entry["products"][product_name] = int(float(val))

            # Other columns
            for col_key, product_name in OTHER_COLS.items():
                idx = col_map.get(col_key)
                if idx is not None and idx < len(row):
                    val = clean_value(row[idx])
                    if val:
                        products_seen.add(product_name)
                        entry["products"][product_name] = val

            # Remarks (last non-None column after the mapped ones, or 'Remarks' col)
            rem_idx = col_map.get("remarks")
            if rem_idx is not None and rem_idx < len(row):
                entry["remarks"] = clean_value(row[rem_idx])

            month_clients[normalise_company(company)] = {
                "raw_name": company,
                **entry,
            }

        monthly_data[month_key] = month_clients
        print(f"  Parsed {sheet_name} ({month_key}): {len(month_clients)} clients")

    return products_seen, monthly_data


def parse_changes(wb):
    """Parse the changes workbook, return list of change records"""
    changes = []

    for sheet_name in wb.sheetnames:
        month_label = sheet_name.strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Header is row 0
        header = [str(c).strip().lower() if c else "" for c in rows[0]]

        def col(name):
            try:
                return header.index(name)
            except ValueError:
                return None

        client_col = 0
        product_col = col("product") or 1
        change_col = col("change") or 2
        date_col = col("date") or 3
        req_col = col("requested by") or 4
        accounts_col = col("accounts informed") or 5
        completed_col = col("change completed by") or 6
        worksheet_col = col("worksheet submitted") or 7
        count_col = col("support count sheet updated")

        for row in rows[1:]:
            if not any(row):
                continue
            client = clean_value(row[client_col] if client_col < len(row) else None)
            if not client or "example" in client.lower():
                continue

            product = clean_value(row[product_col] if product_col < len(row) else None)
            change_desc = clean_value(row[change_col] if change_col < len(row) else None)
            if not change_desc and not product:
                continue

            date_val = row[date_col] if date_col < len(row) else None
            if isinstance(date_val, datetime):
                change_date = date_val.replace(tzinfo=timezone.utc)
            elif date_val:
                try:
                    change_date = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").replace(tzinfo=timezone.utc)
                except ValueError:
                    change_date = datetime.now(timezone.utc)
            else:
                change_date = datetime.now(timezone.utc)

            accounts_raw = clean_value(row[accounts_col] if accounts_col < len(row) else None)
            accounts_informed = bool(accounts_raw and accounts_raw.lower() not in ("no", "n/a", "na", "false"))

            worksheet_raw = clean_value(row[worksheet_col] if worksheet_col < len(row) else None)
            worksheet_submitted = bool(worksheet_raw and worksheet_raw.lower() in ("yes", "y", "true"))

            count_updated_raw = clean_value(row[count_col] if count_col is not None and count_col < len(row) else None)
            profile_updated = bool(count_updated_raw and "yes" in count_updated_raw.lower())

            changes.append({
                "id": str(uuid.uuid4()),
                "raw_client_name": client,
                "client_id": None,  # resolved later
                "product_name": product,
                "product_id": None,  # resolved later
                "change_description": change_desc or product or "",
                "date": change_date.isoformat(),
                "requested_by": clean_value(row[req_col] if req_col < len(row) else None),
                "completed_by": clean_value(row[completed_col] if completed_col < len(row) else None),
                "accounts_informed": accounts_informed,
                "worksheet_submitted": worksheet_submitted,
                "profile_updated": profile_updated,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": "migration",
            })

    print(f"  Parsed {len(changes)} change records")
    return changes


def build_products(products_seen: set) -> list:
    """Build the product catalogue from all products encountered"""
    products = []
    sort_map = {
        "security": 10, "backup": 20, "devices": 30, "onsite": 40,
        "connectivity": 50, "hosting": 60, "office365": 70, "other": 80,
    }
    for i, name in enumerate(sorted(products_seen)):
        cat = PRODUCT_CATEGORIES.get(name, "other")
        products.append({
            "id": str(uuid.uuid4()),
            "name": name,
            "category": cat,
            "unit": PRODUCT_UNITS.get(name, "count"),
            "active": True,
            "sort_order": sort_map.get(cat, 90) + i,
            "unit_cost": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    return products


def resolve_client_ids(monthly_data: dict, existing_clients: list) -> dict:
    """
    Map normalised company names to SynthOps client IDs.
    Returns: { normalised_name: client_id }
    Also prints unresolved names for manual review.
    """
    mapping = {}
    # Build lookup from existing clients
    client_lookup = {}
    for c in existing_clients:
        name = (c.get("name") or "").strip().lower()
        client_lookup[name] = c["id"]
        # Also try without common suffixes
        for suffix in [" ltd", " limited", " plc", " inc", " llc", " solutions", " group"]:
            if name.endswith(suffix):
                client_lookup[name[:-len(suffix)].strip()] = c["id"]

    unresolved = set()
    all_names = set()
    for month_clients in monthly_data.values():
        for norm_name, data in month_clients.items():
            all_names.add((norm_name, data["raw_name"]))

    for norm_name, raw_name in sorted(all_names):
        if norm_name in client_lookup:
            mapping[norm_name] = client_lookup[norm_name]
        else:
            # Try partial matching
            matched = None
            for client_name, client_id in client_lookup.items():
                if norm_name in client_name or client_name in norm_name:
                    matched = client_id
                    break
            if matched:
                mapping[norm_name] = matched
            else:
                unresolved.add(raw_name)
                # Use a placeholder - can be fixed after import
                mapping[norm_name] = f"UNRESOLVED:{raw_name}"

    if unresolved:
        print(f"\n  ⚠️  {len(unresolved)} clients could not be matched to SynthOps client IDs:")
        for name in sorted(unresolved):
            print(f"     - {name}")
        print("  These will be imported with UNRESOLVED: prefix in client_id.")
        print("  You can update them manually in MongoDB after import.\n")

    return mapping


def build_snapshots_and_profiles(monthly_data: dict, client_id_map: dict) -> tuple[list, dict]:
    """Build snapshot records and final profiles (from latest month)"""
    snapshots = []
    # Track latest profile per client
    latest_by_client = {}

    sorted_months = sorted(monthly_data.keys())

    for month in sorted_months:
        month_clients = monthly_data[month]
        for norm_name, data in month_clients.items():
            client_id = client_id_map.get(norm_name, f"UNRESOLVED:{data['raw_name']}")
            snap = {
                "client_id": client_id,
                "month": month,
                "support_type": data["support_type"],
                "remarks": data["remarks"],
                "products": data["products"],
                "snapshot_date": datetime.now(timezone.utc).isoformat(),
            }
            snapshots.append(snap)
            latest_by_client[client_id] = snap

    # Build current profiles from the most recent snapshot per client
    profiles = []
    for client_id, snap in latest_by_client.items():
        profiles.append({
            "id": str(uuid.uuid4()),
            "client_id": client_id,
            "support_type": snap["support_type"],
            "remarks": snap["remarks"],
            "products": snap["products"],
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": "migration",
            "needs_review": False,
        })

    return snapshots, profiles


def resolve_changes(changes: list, client_id_map: dict, product_name_to_id: dict) -> list:
    """Attach client_id and product_id to change records"""
    for change in changes:
        norm = normalise_company(change["raw_client_name"])
        change["client_id"] = client_id_map.get(norm, f"UNRESOLVED:{change['raw_client_name']}")
        if change["product_name"]:
            # Try to match to a product in the catalogue
            pname = change["product_name"].strip().lower()
            for prod_name, prod_id in product_name_to_id.items():
                if pname in prod_name.lower() or prod_name.lower() in pname:
                    change["product_id"] = prod_id
                    break
    return changes


def main():
    parser = argparse.ArgumentParser(description="Import historical SynthOps support data")
    parser.add_argument("--support-count", required=True, help="Path to Monthly_Support_Count.xlsx")
    parser.add_argument("--changes", required=True, help="Path to Monthly_Client_Changes.xlsx")
    parser.add_argument("--url", default="http://localhost:8001", help="SynthOps backend URL")
    parser.add_argument("--token", help="JWT access token (from login)")
    parser.add_argument("--dry-run", action="store_true", help="Print JSON payload, do not POST")
    parser.add_argument("--output", help="Save JSON payload to file")
    args = parser.parse_args()

    print("Loading spreadsheets...")
    sc_wb = load_workbook(args.support_count, read_only=True, data_only=True)
    ch_wb = load_workbook(args.changes, read_only=True, data_only=True)

    print("\nParsing Support Count sheets...")
    products_seen, monthly_data = parse_support_count(sc_wb)

    print("\nParsing Changes sheets...")
    changes = parse_changes(ch_wb)

    print("\nBuilding product catalogue...")
    products = build_products(products_seen)
    product_name_to_id = {p["name"]: p["id"] for p in products}
    print(f"  {len(products)} products found")

    # Try to fetch existing clients from SynthOps to resolve IDs
    existing_clients = []
    if not args.dry_run and args.token:
        try:
            resp = requests.get(
                f"{args.url}/api/clients",
                headers={"Authorization": f"Bearer {args.token}"},
                timeout=10
            )
            if resp.ok:
                existing_clients = resp.json()
                print(f"\nFetched {len(existing_clients)} existing clients from SynthOps")
        except Exception as e:
            print(f"\nCould not fetch clients from SynthOps: {e}")

    print("\nResolving client IDs...")
    client_id_map = resolve_client_ids(monthly_data, existing_clients)

    print("\nBuilding snapshots and profiles...")
    snapshots, profiles = build_snapshots_and_profiles(monthly_data, client_id_map)
    print(f"  {len(snapshots)} snapshots, {len(profiles)} current profiles")

    print("\nResolving change records...")
    changes = resolve_changes(changes, client_id_map, product_name_to_id)

    payload = {
        "products": products,
        "profiles": profiles,
        "snapshots": snapshots,
        "changes": changes,
    }

    print(f"\nPayload summary:")
    print(f"  Products:  {len(products)}")
    print(f"  Profiles:  {len(profiles)}")
    print(f"  Snapshots: {len(snapshots)}")
    print(f"  Changes:   {len(changes)}")

    if args.output:
        with open(args.output, "w") as f:
            json.dump(payload, f, indent=2, default=str)
        print(f"\nPayload saved to {args.output}")

    if args.dry_run:
        print("\nDry run complete. Use --output to save the payload.")
        return

    if not args.token:
        print("\nNo --token provided. Use --dry-run or provide a token to import.")
        return

    print(f"\nPosting to {args.url}/api/support/import ...")
    resp = requests.post(
        f"{args.url}/api/support/import",
        json=payload,
        headers={
            "Authorization": f"Bearer {args.token}",
            "Content-Type": "application/json",
        },
        timeout=120,
    )
    if resp.ok:
        result = resp.json()
        print("\n✅ Import complete:")
        print(f"   Products:  {result.get('products', 0)}")
        print(f"   Profiles:  {result.get('profiles', 0)}")
        print(f"   Snapshots: {result.get('snapshots', 0)}")
        print(f"   Changes:   {result.get('changes', 0)}")
        if result.get("errors"):
            print(f"\n⚠️  {len(result['errors'])} errors:")
            for err in result["errors"][:20]:
                print(f"   - {err}")
    else:
        print(f"\n❌ Import failed: {resp.status_code}")
        print(resp.text[:500])


if __name__ == "__main__":
    main()
